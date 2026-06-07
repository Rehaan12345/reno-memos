"""
ingest.py — Layer 1 (Knowledge Layer)

Loads the CityThread Q1 2026 master workbook into reno.db.

Six sheets are ingested: Master Log, Metrics, Issue Threads, Decisions, Flags,
Category Key. The README and Agent Context sheets are intentionally skipped.

All inserts use INSERT OR REPLACE so re-ingesting an updated spreadsheet is safe.
Absent content is stored as NULL, never invented (see CLAUDE.md). global_id is
the canonical reference across every table.

Usage:
    python ingest.py [path/to/workbook.xlsx]   # default: CityThread_Q1_2026_Master.xlsx
"""

import json
import re
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

WORKBOOK_DEFAULT = "CityThread_Q1_2026_Master.xlsx"
DB_PATH = "reno.db"

# Sheet names as they appear in the workbook.
SHEET_MASTER = "01 · Master Log"
SHEET_METRICS = "02 · Metrics"
SHEET_THREADS = "03 · Issue Threads"
SHEET_DECISIONS = "04 · Decisions"
SHEET_FLAGS = "05 · Flags"
SHEET_CATEGORIES = "07 · Category Key"

GLOBAL_ID_RE = re.compile(r"\b([A-Z]{3}-\d{3})\b")
MONTH_ABBR = {"Jan": "JAN", "Feb": "FEB", "Mar": "MAR", "Apr": "APR"}


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def clean(value):
    """Normalize a cell value to a trimmed string, or None if empty/'None'."""
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() == "none":
        return None
    # The spreadsheet uses "None — informational report only" style strings.
    # Those carry meaning, so keep them; only a bare "None" becomes NULL.
    return s


def to_iso_date(value):
    """Parse the sheet's date forms ('02-Jan-26' or a datetime) to ISO YYYY-MM-DD."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Unparseable — keep the raw string rather than dropping the fact.
    return s


def header_index(ws, header_row=1):
    """Map normalized header label -> column index (1-based) for a sheet."""
    idx = {}
    for c in range(1, ws.max_column + 1):
        label = clean(ws.cell(header_row, c).value)
        if label:
            idx[label] = c
    return idx


def extract_global_ids(text, source_month=None):
    """
    Pull canonical global_ids from a free-text cell.

    Handles the documented Related-Memos formats:
      "002, 004, 007"            -> bare numbers, prefixed with source memo's month
      "FEB-008,MAR-003"          -> already canonical
      "001 (Dec), 003 (parking)" -> strip parentheticals, then as bare numbers
      "All weekly reports"       -> yields nothing
    """
    if not text:
        return []
    # Drop parenthetical notes like "(Dec)" or "(parking)".
    stripped = re.sub(r"\([^)]*\)", " ", str(text))
    ids = []
    # Canonical IDs first.
    for m in GLOBAL_ID_RE.finditer(stripped):
        ids.append(m.group(1))
    # Remove canonical matches so their digits aren't re-counted as bare numbers.
    remainder = GLOBAL_ID_RE.sub(" ", stripped)
    if source_month:
        prefix = MONTH_ABBR.get(source_month)
        if prefix:
            for m in re.finditer(r"\b(\d{1,3})\b", remainder):
                ids.append(f"{prefix}-{int(m.group(1)):03d}")
    # De-duplicate, preserve order.
    seen, out = set(), []
    for i in ids:
        if i not in seen:
            seen.add(i)
            out.append(i)
    return out


# --------------------------------------------------------------------------- #
# Schema
# --------------------------------------------------------------------------- #
def create_schema(conn):
    conn.executescript(
        """
        DROP TABLE IF EXISTS memos;
        DROP TABLE IF EXISTS relationships;
        DROP TABLE IF EXISTS threads;
        DROP TABLE IF EXISTS flags;
        DROP TABLE IF EXISTS decisions;
        DROP TABLE IF EXISTS metrics;
        DROP TABLE IF EXISTS categories;
        DROP TABLE IF EXISTS memos_fts;

        CREATE TABLE memos (
            global_id      TEXT PRIMARY KEY,
            month          TEXT,
            memo_id        INTEGER,
            date_published TEXT,
            title          TEXT,
            department     TEXT,
            authors        TEXT,
            category       TEXT,
            subcategory    TEXT,
            tldr           TEXT,
            key_stats      TEXT,
            key_decisions  TEXT,
            action_items   TEXT,
            keywords       TEXT,
            source_url     TEXT,
            search_text    TEXT
        );

        CREATE TABLE relationships (
            source_id TEXT,
            target_id TEXT,
            rel_type  TEXT,
            PRIMARY KEY (source_id, target_id, rel_type)
        );

        CREATE TABLE threads (
            name       TEXT PRIMARY KEY,
            status     TEXT,
            memo_ids   TEXT,   -- JSON array of global_ids
            first_memo TEXT,
            latest_memo TEXT,
            jan        TEXT,
            feb        TEXT,
            mar        TEXT,
            apr        TEXT,
            key_signal TEXT
        );

        CREATE TABLE flags (
            flag        TEXT PRIMARY KEY,
            priority    TEXT,
            signal      TEXT,
            source_memos TEXT,   -- JSON array of global_ids
            source_raw  TEXT,    -- original cell text (preserves "All weekly reports")
            status      TEXT,
            next_step   TEXT
        );

        CREATE TABLE decisions (
            global_id  TEXT,
            date       TEXT,
            month      TEXT,
            department TEXT,
            decision   TEXT,
            made_by    TEXT,
            category   TEXT,
            impact     TEXT
        );

        CREATE TABLE metrics (
            series_name      TEXT PRIMARY KEY,
            unit             TEXT,
            description      TEXT,
            source_memo_ids  TEXT,   -- JSON array of global_ids
            periods          TEXT,   -- JSON array
            data_values      TEXT    -- JSON array (parallel to periods)
        );

        CREATE TABLE categories (
            category        TEXT PRIMARY KEY,
            original_labels TEXT,
            description     TEXT,
            count           INTEGER
        );

        CREATE VIRTUAL TABLE memos_fts USING fts5(
            global_id UNINDEXED,
            search_text
        );
        """
    )


# --------------------------------------------------------------------------- #
# Per-sheet ingestion
# --------------------------------------------------------------------------- #
def ingest_memos(conn, ws):
    h = header_index(ws)

    def col(label):
        return h.get(label)

    rows = 0
    for r in range(2, ws.max_row + 1):
        gid = clean(ws.cell(r, col("Global ID")).value)
        if not gid:
            continue

        def g(label):
            ci = col(label)
            return clean(ws.cell(r, ci).value) if ci else None

        memo_id = g("Memo ID")
        record = {
            "global_id": gid,
            "month": g("Month"),
            "memo_id": int(memo_id) if memo_id and memo_id.isdigit() else None,
            "date_published": to_iso_date(ws.cell(r, col("Date Published")).value),
            "title": g("Title"),
            "department": g("Department"),
            "authors": g("Author(s)"),
            "category": g("Category (Normalised)"),
            "subcategory": g("Sub-category"),
            "tldr": g("TL;DR"),
            "key_stats": g("Key Stats"),
            "key_decisions": g("Key Decisions"),
            "action_items": g("Action Items / Deadlines"),
            "keywords": g("Keywords / Tags"),
            "source_url": g("Source URL"),
        }
        # search_text concatenates all text fields for FTS indexing.
        searchable = [
            record["title"], record["department"], record["authors"],
            record["category"], record["subcategory"], record["tldr"],
            record["key_stats"], record["key_decisions"],
            record["action_items"], record["keywords"],
        ]
        record["search_text"] = " ".join(x for x in searchable if x)

        conn.execute(
            """INSERT OR REPLACE INTO memos
               (global_id, month, memo_id, date_published, title, department,
                authors, category, subcategory, tldr, key_stats, key_decisions,
                action_items, keywords, source_url, search_text)
               VALUES (:global_id, :month, :memo_id, :date_published, :title,
                :department, :authors, :category, :subcategory, :tldr,
                :key_stats, :key_decisions, :action_items, :keywords,
                :source_url, :search_text)""",
            record,
        )

        # Relationships from the Related Memos column (default rel_type=references).
        related_ci = col("Related Memos")
        related_raw = clean(ws.cell(r, related_ci).value) if related_ci else None
        for target in extract_global_ids(related_raw, source_month=record["month"]):
            if target == gid:
                continue
            conn.execute(
                "INSERT OR REPLACE INTO relationships (source_id, target_id, rel_type)"
                " VALUES (?, ?, 'references')",
                (gid, target),
            )
        rows += 1
    return rows


def ingest_threads(conn, ws):
    h = header_index(ws)
    rows = 0
    for r in range(2, ws.max_row + 1):
        name = clean(ws.cell(r, h["Thread"]).value)
        if not name:
            continue
        memo_ids = extract_global_ids(clean(ws.cell(r, h["Memos (Global IDs)"]).value))
        conn.execute(
            """INSERT OR REPLACE INTO threads
               (name, status, memo_ids, first_memo, latest_memo, jan, feb, mar, apr, key_signal)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                name,
                clean(ws.cell(r, h["Status"]).value),
                json.dumps(memo_ids),
                clean(ws.cell(r, h["First Memo"]).value),
                clean(ws.cell(r, h["Latest Memo"]).value),
                clean(ws.cell(r, h["Jan"]).value),
                clean(ws.cell(r, h["Feb"]).value),
                clean(ws.cell(r, h["Mar"]).value),
                clean(ws.cell(r, h["Apr"]).value),
                clean(ws.cell(r, h["Key Signal / Takeaway"]).value),
            ),
        )
        rows += 1
    return rows


def ingest_decisions(conn, ws):
    h = header_index(ws)
    rows = 0
    for r in range(2, ws.max_row + 1):
        gid = clean(ws.cell(r, h["Memo ID"]).value)
        if not gid:
            continue
        conn.execute(
            """INSERT OR REPLACE INTO decisions
               (global_id, date, month, department, decision, made_by, category, impact)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                gid,
                to_iso_date(ws.cell(r, h["Date"]).value),
                clean(ws.cell(r, h["Month"]).value),
                clean(ws.cell(r, h["Department"]).value),
                clean(ws.cell(r, h["Decision"]).value),
                clean(ws.cell(r, h["Made By"]).value),
                clean(ws.cell(r, h["Category"]).value),
                clean(ws.cell(r, h["Impact"]).value),
            ),
        )
        rows += 1
    return rows


def ingest_flags(conn, ws):
    h = header_index(ws)
    rows = 0
    for r in range(2, ws.max_row + 1):
        flag = clean(ws.cell(r, h["Flag"]).value)
        if not flag:
            continue
        source_raw = clean(ws.cell(r, h["Source Memos"]).value)
        conn.execute(
            """INSERT OR REPLACE INTO flags
               (flag, priority, signal, source_memos, source_raw, status, next_step)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                flag,
                clean(ws.cell(r, h["Priority"]).value),
                clean(ws.cell(r, h["Signal"]).value),
                json.dumps(extract_global_ids(source_raw)),
                source_raw,
                clean(ws.cell(r, h["Status"]).value),
                clean(ws.cell(r, h["Next Step"]).value),
            ),
        )
        rows += 1
    return rows


def ingest_categories(conn, ws):
    h = header_index(ws)
    rows = 0
    for r in range(2, ws.max_row + 1):
        cat = clean(ws.cell(r, h["Normalised Category"]).value)
        if not cat:
            continue
        count = clean(ws.cell(r, h["Count in Q1"]).value)
        conn.execute(
            """INSERT OR REPLACE INTO categories
               (category, original_labels, description, count) VALUES (?, ?, ?, ?)""",
            (
                cat,
                clean(ws.cell(r, h["Original Labels (merged)"]).value),
                clean(ws.cell(r, h["Description"]).value),
                int(count) if count and count.isdigit() else None,
            ),
        )
        rows += 1
    return rows


def ingest_metrics(conn, ws):
    """
    The metrics sheet is not tabular. Each series is a 4-row block:

        <Series Name>            | ... | Unit: <unit>
        Source memos: <ids>   |   <description>
        Period | p1 | p2 | ...
        Value  | v1 | v2 | ...

    Parse by scanning for these sentinel patterns, not by column position.
    """
    rows = 0
    r = 1
    max_r = ws.max_row
    while r <= max_r:
        c1 = clean(ws.cell(r, 1).value)
        # A series header: non-empty col1 that is not a sentinel keyword, and the
        # row carries a "Unit:" marker somewhere across the row.
        if c1 and not c1.startswith(("Source memos", "Period", "Value")):
            unit = None
            for c in range(2, ws.max_column + 1):
                cell = clean(ws.cell(r, c).value)
                if cell and cell.startswith("Unit:"):
                    unit = cell.split(":", 1)[1].strip()
                    break
            if unit is not None:
                series_name = c1
                source_ids, description = [], None
                periods, values = [], []
                # Scan the block until the next header or blank gap.
                rr = r + 1
                while rr <= max_r:
                    a = clean(ws.cell(rr, 1).value)
                    if a is None:
                        rr += 1
                        # Stop if two-cell lookahead shows a new header block.
                        nxt = clean(ws.cell(rr, 1).value) if rr <= max_r else None
                        if nxt and not nxt.startswith(("Source memos", "Period", "Value")):
                            break
                        continue
                    if a.startswith("Source memos"):
                        # Format: "Source memos: A, B  |  <description>"
                        body = a.split(":", 1)[1] if ":" in a else a
                        parts = body.split("|", 1)
                        source_ids = extract_global_ids(parts[0])
                        description = parts[1].strip() if len(parts) > 1 else None
                    elif a == "Period":
                        periods = [
                            clean(ws.cell(rr, c).value)
                            for c in range(2, ws.max_column + 1)
                            if clean(ws.cell(rr, c).value) is not None
                        ]
                    elif a == "Value":
                        for c in range(2, ws.max_column + 1):
                            v = ws.cell(rr, c).value
                            if v is not None:
                                values.append(v)
                        # Value row ends the block.
                        rr += 1
                        break
                    elif not a.startswith(("Source memos", "Period", "Value")):
                        # Next series header reached.
                        break
                    rr += 1

                conn.execute(
                    """INSERT OR REPLACE INTO metrics
                       (series_name, unit, description, source_memo_ids, periods, data_values)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (
                        series_name,
                        unit,
                        description,
                        json.dumps(source_ids),
                        json.dumps(periods),
                        json.dumps(values),
                    ),
                )
                rows += 1
                r = rr
                continue
        r += 1
    return rows


def build_fts(conn):
    conn.execute("DELETE FROM memos_fts")
    conn.execute(
        "INSERT INTO memos_fts (global_id, search_text) "
        "SELECT global_id, search_text FROM memos"
    )


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
EXPECTED = {
    "memos": 77,
    "metrics": 18,
    "threads": 12,
    "decisions": 14,
    "flags": 10,
    "categories": 18,
}


def main():
    workbook = sys.argv[1] if len(sys.argv) > 1 else WORKBOOK_DEFAULT
    if not Path(workbook).exists():
        sys.exit(f"Workbook not found: {workbook}")

    wb = openpyxl.load_workbook(workbook, data_only=True)
    conn = sqlite3.connect(DB_PATH)
    try:
        create_schema(conn)
        counts = {
            "memos": ingest_memos(conn, wb[SHEET_MASTER]),
            "metrics": ingest_metrics(conn, wb[SHEET_METRICS]),
            "threads": ingest_threads(conn, wb[SHEET_THREADS]),
            "decisions": ingest_decisions(conn, wb[SHEET_DECISIONS]),
            "flags": ingest_flags(conn, wb[SHEET_FLAGS]),
            "categories": ingest_categories(conn, wb[SHEET_CATEGORIES]),
        }
        build_fts(conn)
        rel_count = conn.execute("SELECT COUNT(*) FROM relationships").fetchone()[0]
        conn.commit()
    finally:
        conn.close()

    print(f"Ingested {workbook} -> {DB_PATH}")
    ok = True
    for table, n in counts.items():
        expect = EXPECTED[table]
        mark = "ok " if n == expect else "DIFF"
        if n != expect:
            ok = False
        print(f"  {mark} {table:<11} {n:>3}  (expected {expect})")
    print(f"  --  relationships {rel_count:>3}  (edges)")
    if not ok:
        print("\nWARNING: row counts differ from spreadsheet expectations.")
        sys.exit(1)
    print("\nAll row counts match. Knowledge layer ready.")


if __name__ == "__main__":
    main()
